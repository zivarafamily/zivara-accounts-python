from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from app.database import SessionLocal
from app.models import CashBookEntry, Expense, LLP, LLPPartner, LLPPayable, NeoRevenue, UploadedBill, User
from app.security import hash_password
from app.services.payables import calculate_amounts, normalize_tds_section


def test_auth(client):
    res = client.post("/auth/login", json={"username": "admin", "password": "ChangeMe123!"})
    assert res.status_code == 200
    assert res.json()["access_token"]
    assert res.json()["llps"][0]["llpId"] == "LLP001"
    assert "pan" in res.json()["llps"][0]
    assert "address" in res.json()["llps"][0]


def test_auth_me_works_with_token(client):
    login = client.post("/auth/login", json={"username": "admin", "password": "ChangeMe123!"})
    token = login.json()["access_token"]
    res = client.get("/auth/me", headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 200
    assert res.json()["user"] == "admin"


def test_admin_can_save_multiple_llps_without_short_code(client, auth_headers):
    first = client.post("/llps", headers=auth_headers, json={"LLPName": "Alpha Family Office LLP"})
    second = client.post("/llps", headers=auth_headers, json={"LLPName": "Alpha Finance LLP"})
    assert first.status_code == 200
    assert second.status_code == 200

    rows = client.get("/llps", headers=auth_headers).json()["data"]
    created = [row for row in rows if row["LLPName"].startswith("Alpha")]
    assert len(created) == 2
    assert all(row["ShortCode"] for row in created)
    assert len({row["ShortCode"] for row in created}) == 2


def test_llp_save_rejects_duplicate_name_and_short_code(client, auth_headers):
    first = client.post("/llps", headers=auth_headers, json={"LLPName": "Beta LLP", "ShortCode": "BETA"})
    assert first.status_code == 200

    duplicate_name = client.post("/llps", headers=auth_headers, json={"LLPName": " beta llp ", "ShortCode": "BETA2"})
    duplicate_code = client.post("/llps", headers=auth_headers, json={"LLPName": "Beta Two LLP", "ShortCode": "beta"})
    assert duplicate_name.status_code == 409
    assert duplicate_name.json()["detail"] == "LLP name already exists"
    assert duplicate_code.status_code == 409
    assert duplicate_code.json()["detail"] == "Short Code already exists"


def test_payable_calculations():
    taxable, gst, gross, rate, tds, net = calculate_amounts({"TaxableAmount": "1000", "GSTAmount": "180", "TDSRate": "10"})
    assert taxable == Decimal("1000.00")
    assert gst == Decimal("180.00")
    assert gross == Decimal("1180.00")
    assert tds == Decimal("100.00")
    assert net == Decimal("1080.00")


def test_legacy_tds_sections_normalize_to_section_393():
    assert normalize_tds_section("194C", "2", "Travel Agency") == "393(1)-6(i)-2"
    assert normalize_tds_section("194J", "10", "Software") == "393(1)-6(iii)-10"
    assert normalize_tds_section("194Q", "0.1", "Office Purchase") == "393(1)-8(ii)"


def test_duplicate_payable_rejection(client, auth_headers):
    v = client.post("/vendors", headers=auth_headers, json={"VendorName": "ABC Co"})
    vendor_id = v.json()["data"]["VendorID"]
    payload = {"VendorID": vendor_id, "VendorName": "ABC Co", "BillNo": "INV-1", "BillDate": "2026-06-10", "TaxableAmount": "1000", "GSTAmount": "180", "TDSRate": "10"}
    assert client.post("/payables", headers=auth_headers, json=payload).status_code == 200
    assert client.post("/payables", headers=auth_headers, json=payload).status_code == 409


def test_vendor_delete_is_blocked_when_payables_exist(client, auth_headers):
    vendor = client.post("/vendors", headers=auth_headers, json={"VendorName": "Delete Block Vendor"})
    assert vendor.status_code == 200
    vendor_id = vendor.json()["data"]["VendorID"]
    payable = client.post(
        "/payables",
        headers=auth_headers,
        json={"VendorID": vendor_id, "VendorName": "Delete Block Vendor", "BillNo": "DEL-1", "BillDate": "2026-06-10", "GrossAmount": "1000"},
    )
    assert payable.status_code == 200

    deleted = client.delete(f"/vendors/{vendor_id}", headers=auth_headers)
    assert deleted.status_code == 409
    assert "payables" in deleted.json()["detail"]


def test_unused_vendor_can_be_deleted(client, auth_headers):
    vendor = client.post("/vendors", headers=auth_headers, json={"VendorName": "Unused Delete Vendor"})
    assert vendor.status_code == 200
    vendor_id = vendor.json()["data"]["VendorID"]

    deleted = client.delete(f"/vendors/{vendor_id}", headers=auth_headers)
    assert deleted.status_code == 200
    rows = client.get("/vendors", headers=auth_headers).json()["data"]
    assert all(row["VendorID"] != vendor_id for row in rows)


def test_payable_tds_section_is_normalized_and_editable(client, auth_headers):
    created = client.post(
        "/payables",
        headers=auth_headers,
        json={
            "VendorName": "Travel Co",
            "VendorCategory": "Travel Agency",
            "BillNo": "TDS-1",
            "BillDate": "2026-05-18",
            "GrossAmount": "7167",
            "TDSSection": "194C",
            "TDSAmount": "2",
        },
    )
    assert created.status_code == 200
    payable_id = created.json()["data"]["PayableID"]

    rows = client.get("/payables", headers=auth_headers).json()["data"]
    row = next(r for r in rows if r["PayableID"] == payable_id)
    assert row["TDSSection"] == "393(1)-6(i)-2"
    assert row["TDSSectionRaw"] == "393(1)-6(i)-2"

    updated = client.put(
        f"/payables/{payable_id}",
        headers=auth_headers,
        json={"TDSSection": "393(1)-8(ii)", "TDSRate": "0.1", "TDSAmount": "7.17"},
    )
    assert updated.status_code == 200
    row = next(r for r in client.get("/payables", headers=auth_headers).json()["data"] if r["PayableID"] == payable_id)
    assert row["TDSSection"] == "393(1)-8(ii)"
    assert row["TDSRate"] == 0.1


def test_payable_update_persists_edited_fields(client, auth_headers):
    created = client.post(
        "/payables",
        headers=auth_headers,
        json={
            "VendorName": "Editable Vendor",
            "VendorCategory": "Travel Agency",
            "BillNo": "EDIT-1",
            "BillDate": "2026-05-18",
            "TaxableAmount": "1000",
            "GSTAmount": "180",
            "TDSRate": "2",
            "Notes": "before",
        },
    )
    assert created.status_code == 200
    payable_id = created.json()["data"]["PayableID"]

    updated = client.put(
        f"/payables/{payable_id}",
        headers=auth_headers,
        json={
            "VendorName": "Editable Vendor Updated",
            "VendorCategory": "Office Purchase",
            "BillNo": "EDIT-2",
            "BillDate": "2026-05-20",
            "TaxableAmount": "2000",
            "GSTAmount": "360",
            "GrossAmount": "2360",
            "TDSRate": "0",
            "TDSAmount": "0",
            "PaymentMode": "NEFT",
            "BankAccount": "",
            "ReferenceNo": "",
            "Notes": "",
        },
    )
    assert updated.status_code == 200
    data = updated.json()["data"]
    assert data["VendorName"] == "Editable Vendor Updated"
    assert data["VendorCategory"] == "Office Purchase"
    assert data["BillNo"] == "EDIT-2"
    assert data["BillDate"] == "2026-05-20"
    assert data["GrossAmount"] == 2360.0
    assert data["TDSAmount"] == 0.0
    assert data["BankAccount"] == ""
    assert data["Notes"] == ""


def test_vendor_payables_can_be_marked_paid_in_batch(client, auth_headers):
    ids = []
    for bill_no, amount in [("BATCH-1", "1000"), ("BATCH-2", "2500")]:
        created = client.post(
            "/payables",
            headers=auth_headers,
            json={
                "VendorName": "Batch Vendor",
                "VendorCategory": "Travel Agency",
                "BillNo": bill_no,
                "BillDate": "2026-05-18",
                "GrossAmount": amount,
            },
        )
        assert created.status_code == 200
        ids.append(created.json()["data"]["PayableID"])

    paid = client.post(
        "/payables/batch-payment",
        headers=auth_headers,
        json={
            "PayableIDs": ids,
            "PaymentDate": "2026-05-20",
            "PaymentMode": "NEFT",
            "BankAccount": "HDFC",
            "ReferenceNo": "UTR-BATCH",
        },
    )
    assert paid.status_code == 200
    body = paid.json()
    assert body["paidCount"] == 2
    assert body["paidAmount"] == 3500.0

    rows = [r for r in client.get("/payables", headers=auth_headers).json()["data"] if r["PayableID"] in ids]
    assert {r["Status"] for r in rows} == {"Paid"}
    assert {r["ReferenceNo"] for r in rows} == {"UTR-BATCH"}


def test_vendor_ledger_returns_headers_rows_and_summary(client, auth_headers):
    created = client.post(
        "/payables",
        headers=auth_headers,
        json={
            "VendorName": "Ledger Vendor",
            "VendorCategory": "Office Purchase",
            "BillNo": "LED-1",
            "BillDate": "2026-06-10",
            "GrossAmount": "1000",
            "TDSAmount": "100",
            "PaidAmount": "400",
            "PaymentDate": "2026-06-15",
            "ReferenceNo": "UTR-LED",
        },
    )
    assert created.status_code == 200
    res = client.get("/reports/vendor-ledger", headers=auth_headers, params={"vendor": "Ledger Vendor"})
    assert res.status_code == 200
    body = res.json()
    assert "Date" in body["headers"]
    assert len(body["data"]) == 2
    assert body["summary"]["Debit"] == 900.0
    assert body["summary"]["Credit"] == 400.0
    assert body["summary"]["TDSAmount"] == 100.0
    assert body["summary"]["ClosingBalance"] == 500.0


def test_expense_duplicate_warning(client, auth_headers):
    payload = {"Date": "2026-06-10", "Amount": "500", "PaidBy": "Dinu", "Description": "Cab"}
    assert client.post("/expenses", headers=auth_headers, json=payload).status_code == 200
    res = client.post("/expenses", headers=auth_headers, json=payload)
    assert res.status_code == 200
    assert res.json()["duplicateWarning"] is True


def test_expense_decimal_gst_and_billing_month_update(client, auth_headers):
    created = client.post(
        "/expenses",
        headers=auth_headers,
        json={
            "Date": "2026-06-10",
            "TaxableValue": "100.50",
            "CGSTAmount": "9.04",
            "SGSTAmount": "9.05",
            "Amount": "118.59",
            "PaidBy": "Dinu",
            "BillingMonth": "Jun-2026",
        },
    )
    assert created.status_code == 200
    expense_id = created.json()["data"]["ExpenseID"]
    row = next(r for r in client.get("/expenses", headers=auth_headers).json()["data"] if r["ExpenseID"] == expense_id)
    assert row["GSTAmount"] == 18.09
    assert row["BillingMonth"] == "Jun-2026"

    updated = client.put(
        f"/expenses/{expense_id}",
        headers=auth_headers,
        json={"CGSTAmount": "1.25", "SGSTAmount": "1.75", "IGSTAmount": "0", "BillingMonth": "Jul-2026"},
    )
    assert updated.status_code == 200
    row = next(r for r in client.get("/expenses", headers=auth_headers).json()["data"] if r["ExpenseID"] == expense_id)
    assert row["GSTAmount"] == 3.0
    assert row["BillingMonth"] == "Jul-2026"


def test_expense_approval_and_petty_cash_reimbursement(client, auth_headers):
    created = client.post("/expenses", headers=auth_headers, json={"Date": "2026-06-10", "Amount": "750", "PaidBy": "Dinu", "Description": "Fuel"}).json()
    expense_id = created["data"]["ExpenseID"]
    assert client.post(f"/expenses/{expense_id}/approve", headers=auth_headers, json={}).status_code == 200
    res = client.post(f"/expenses/{expense_id}/reimburse", headers=auth_headers, json={"ReimburseMode": "Petty Cash", "ReimburseAccount": "Petty Cash"})
    assert res.status_code == 200
    cash = client.get("/cash-book", headers=auth_headers).json()["data"]
    assert len(cash) == 1
    assert cash[0]["AmountOut"] == 750.0


def test_dashboard_and_reports(client, auth_headers):
    client.post("/receipts", headers=auth_headers, json={"Date": "2026-06-10", "AmountReceived": "1000"})
    dash = client.get("/reports/dashboard", headers=auth_headers).json()["summary"]
    assert dash["receipts_total"] == 1000.0
    assert client.get("/reports/reconciliation", headers=auth_headers).status_code == 200
    assert client.get("/reports/ca-tds", headers=auth_headers).status_code == 200


def test_neo_invoice_create_list_update(client, auth_headers):
    payload = {
        "InvoiceNo": "ZivNeo001",
        "InvoiceDate": "2026-06-13",
        "BillingMonth": "Jun-2026",
        "InvoiceType": "Professional Fees",
        "GSTMode": "With GST",
        "IsProforma": "No",
        "Particulars": "Professional Fees",
        "TaxableAmount": "1000",
        "GSTRate": "18",
        "GSTAmount": "180",
        "Amount": "1180",
        "TDSRate": "10",
        "TDSAmount": "100",
        "NetPayable": "1080",
    }
    created = client.post("/neo-invoices", headers=auth_headers, json=payload)
    assert created.status_code == 200
    body = created.json()["data"]
    assert body["InvoiceTitle"] == "Tax Invoice"
    assert body["Amount"] == 1180.0

    rows = client.get("/neo-invoices", headers=auth_headers).json()["data"]
    assert len(rows) == 1
    invoice_id = rows[0]["NeoInvoiceID"]

    updated = client.put(f"/neo-invoices/{invoice_id}", headers=auth_headers, json={"Status": "Sent"})
    assert updated.status_code == 200
    assert updated.json()["data"]["Status"] == "Sent"


def test_admin_neo_invoice_list_shows_all_llps(client, auth_headers):
    db = SessionLocal()
    db.add(LLP(id="LLP002", llp_name="Second LLP", short_code="SECOND", gstin="", pan="", address="", status="Active"))
    db.commit()
    db.close()

    base_payload = {
        "InvoiceDate": "2026-06-13",
        "BillingMonth": "Jun-2026",
        "InvoiceType": "Professional Fees",
        "GSTMode": "With GST",
        "IsProforma": "No",
        "Particulars": "Professional Fees",
        "TaxableAmount": "1000",
        "GSTRate": "18",
        "GSTAmount": "180",
        "Amount": "1180",
        "TDSRate": "10",
        "TDSAmount": "100",
        "NetPayable": "1080",
    }
    first_headers = {**auth_headers, "X-LLP-ID": "LLP001"}
    second_headers = {**auth_headers, "X-LLP-ID": "LLP002"}

    first = client.post("/neo-invoices", headers=first_headers, json={**base_payload, "InvoiceNo": "ZivNeoLLP001"})
    second = client.post("/neo-invoices", headers=second_headers, json={**base_payload, "InvoiceNo": "ZivNeoLLP002"})
    assert first.status_code == 200
    assert second.status_code == 200

    rows = client.get("/neo-invoices", headers=auth_headers).json()["data"]
    invoice_numbers = {row["InvoiceNo"] for row in rows}
    assert {"ZivNeoLLP001", "ZivNeoLLP002"} <= invoice_numbers


def test_partner_llp_name_resolves_to_llp_id(client, auth_headers):
    created = client.post(
        "/partners",
        headers=auth_headers,
        json={"PartnerName": "Partner LLP Test", "LLPName": "Zivara Family Office LLP"},
    )
    assert created.status_code == 200

    rows = client.get("/partners", headers=auth_headers).json()["data"]
    row = next(r for r in rows if r["PartnerName"] == "Partner LLP Test")
    assert row["LLPID"] == "LLP001"
    assert row["LLPName"] == "Zivara Family Office LLP"

    updated = client.put(
        f"/partners/{row['PartnerID']}",
        headers=auth_headers,
        json={"LLPName": ""},
    )
    assert updated.status_code == 200
    rows = client.get("/partners", headers=auth_headers).json()["data"]
    row = next(r for r in rows if r["PartnerName"] == "Partner LLP Test")
    assert row["LLPID"] == ""
    assert row["LLPName"] == ""


def test_clients_enrich_neo_revenue_report(client, auth_headers):
    client_payload = {
        "ClientName": "Client A",
        "PAN": "ABCDE1234F",
        "RMName": "Manugopal",
        "PartnerName": "Manugopal",
        "LLPName": "Zivara Family Office LLP",
        "FamilyName": "Family A",
        "SuperFamilyName": "Super A",
    }
    assert client.post("/clients", headers=auth_headers, json=client_payload).status_code == 200
    revenue_payload = {
        "PAN": "ABCDE1234F",
        "ClientName": "Old Client A",
        "RevenueMonth": "Jun-2026",
        "RevenueAmount": "1000",
        "InvestmentAmount": "50000",
        "YTDValue": "50000",
        "IncomeType": "ARR",
    }
    created = client.post("/neo-revenue", headers=auth_headers, json=revenue_payload)
    assert created.status_code == 200
    data = created.json()["data"]
    assert data["ClientName"] == "Client A"
    assert data["FamilyName"] == "Family A"

    report = client.get("/neo-revenue/report", headers=auth_headers).json()
    assert report["kpis"]["TotalRevenue"] == 1000.0
    assert report["clientWise"][0]["FamilyName"] == "Family A"


def test_neo_revenue_batch_accepts_neo_date_and_short_month(client, auth_headers):
    payload = {
        "rows": [{
            "PAN": "ABTPS5785N",
            "ClientName": "NIKHIL SUNIL",
            "PartnerName": "Manugopal A K",
            "TransactionDate": "30-Apr-26",
            "Product": "PMS",
            "TransactionType": "ARR",
            "SchemeName": "Neo Multi-Asset Moderate Strategy",
            "InvestmentAmount": "57762450.6",
            "RevenueMonth": "Apr-26",
            "RevenueAmount": "1000",
            "IncomeType": "ARR",
        }],
        "statementRef": "Neo Apr-26",
    }
    res = client.post("/neo-revenue/batch", headers=auth_headers, json=payload)
    assert res.status_code == 200
    assert res.json()["saved"] == 1

    rows = client.get("/neo-revenue?search=ABTPS5785N", headers=auth_headers).json()["data"]
    assert rows[0]["TransactionDate"] == "2026-04-30"
    assert rows[0]["RevenueMonth"] == "Apr-26"


def test_partner_neo_revenue_is_scoped_to_own_llp_and_partner(client):
    db = SessionLocal()
    try:
        other_llp = LLP(id="LLP002", llp_name="Other LLP", short_code="OTHER", gstin="", pan="", address="", status="Active")
        partner_user = User(
            id="USR_PARTNER",
            name="Manugopal A K",
            email="partner@zivara.local",
            username="partner",
            password_hash=hash_password("ChangeMe123!"),
            role="partner",
            allowed_modules="neorevenue,neoinvoices",
            status="Active",
        )
        db.add_all([other_llp, partner_user])
        db.flush()
        db.add(LLPPartner(id="MAP_PARTNER", llp_id="LLP001", user_id=partner_user.id, role="partner", percentage="", allowed_modules="neorevenue,neoinvoices", status="Active"))
        db.add_all([
            NeoRevenue(id="REV_OWN", client_name="Own Client", partner_name="Manugopal A K", llp_name="Zivara Family Office LLP", revenue_month="Apr-2026", revenue_amount=Decimal("1000.00")),
            NeoRevenue(id="REV_OTHER_PARTNER", client_name="Other Partner Client", partner_name="Other Partner", llp_name="Zivara Family Office LLP", revenue_month="Apr-2026", revenue_amount=Decimal("2000.00")),
            NeoRevenue(id="REV_OTHER_LLP", client_name="Other LLP Client", partner_name="Manugopal A K", llp_name="Other LLP", revenue_month="Apr-2026", revenue_amount=Decimal("3000.00")),
        ])
        db.commit()
    finally:
        db.close()

    login = client.post("/auth/login", json={"username": "partner", "password": "ChangeMe123!"})
    token = login.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}", "X-LLP-ID": "LLP001"}

    rows = client.get(
        "/neo-revenue?partnerName=Other%20Partner&llpName=Other%20LLP",
        headers=headers,
    ).json()
    assert rows["total"] == 1
    assert rows["data"][0]["RevenueID"] == "REV_OWN"

    meta = client.get("/neo-revenue/meta?partnerName=Other%20Partner&llpName=Other%20LLP", headers=headers).json()
    assert meta["totalRows"] == 1
    assert meta["partners"] == ["Manugopal A K"]

    report = client.get("/neo-revenue/report?partnerName=Other%20Partner&llpName=Other%20LLP", headers=headers).json()
    assert report["kpis"]["TotalRevenue"] == 1000.0
    assert len(report["clientWise"]) == 1

    forbidden = client.get("/neo-revenue", headers={"Authorization": f"Bearer {token}", "X-LLP-ID": "LLP002"})
    assert forbidden.status_code == 403


def test_admin_can_delete_neo_revenue_month(client, auth_headers):
    db = SessionLocal()
    try:
        db.add_all([
            NeoRevenue(id="REV_APR_1", client_name="April One", revenue_month="Apr-2026", revenue_amount=Decimal("100.00")),
            NeoRevenue(id="REV_APR_2", client_name="April Two", revenue_month="Apr-2026", revenue_amount=Decimal("200.00")),
            NeoRevenue(id="REV_MAY_1", client_name="May One", revenue_month="May-2026", revenue_amount=Decimal("300.00")),
        ])
        db.commit()
    finally:
        db.close()

    res = client.delete("/neo-revenue/month/Apr-2026", headers=auth_headers)
    assert res.status_code == 200
    assert res.json()["deleted"] == 2

    rows = client.get("/neo-revenue", headers=auth_headers).json()["data"]
    assert [row["RevenueID"] for row in rows] == ["REV_MAY_1"]


def test_partner_cannot_delete_neo_revenue_month(client):
    db = SessionLocal()
    try:
        partner_user = User(
            id="USR_PARTNER_DELETE",
            name="Partner Delete",
            email="partner-delete@zivara.local",
            username="partner-delete",
            password_hash=hash_password("ChangeMe123!"),
            role="partner",
            allowed_modules="neorevenue",
            status="Active",
        )
        db.add(partner_user)
        db.flush()
        db.add(LLPPartner(id="MAP_PARTNER_DELETE", llp_id="LLP001", user_id=partner_user.id, role="partner", percentage="", allowed_modules="neorevenue", status="Active"))
        db.add(NeoRevenue(id="REV_DENIED_DELETE", client_name="Protected", revenue_month="Apr-2026", revenue_amount=Decimal("100.00")))
        db.commit()
    finally:
        db.close()

    login = client.post("/auth/login", json={"username": "partner-delete", "password": "ChangeMe123!"})
    token = login.json()["access_token"]
    res = client.delete("/neo-revenue/month/Apr-2026", headers={"Authorization": f"Bearer {token}", "X-LLP-ID": "LLP001"})
    assert res.status_code == 403

    db = SessionLocal()
    try:
        assert db.get(NeoRevenue, "REV_DENIED_DELETE") is not None
    finally:
        db.close()


def test_llp_scoped_create_requires_x_llp_id(client):
    login = client.post("/auth/login", json={"username": "admin", "password": "ChangeMe123!"})
    token = login.json()["access_token"]
    res = client.post(
        "/expenses",
        headers={"Authorization": f"Bearer {token}"},
        json={"Date": "2026-06-10", "Amount": "100", "Description": "No LLP"},
    )
    assert res.status_code == 400


def test_llp_scoped_create_succeeds_with_allowed_llp(client, auth_headers):
    res = client.post(
        "/expenses",
        headers=auth_headers,
        json={"Date": "2026-06-10", "Amount": "100", "Description": "With LLP"},
    )
    assert res.status_code == 200


def test_admin_can_create_user_and_assign_llp_access(client, auth_headers):
    created = client.post(
        "/users",
        headers=auth_headers,
        json={
            "Name": "Test User",
            "Username": "testuser@zivara.local",
            "Email": "testuser@zivara.local",
            "Password": "ChangeMe123!",
            "Role": "viewer",
            "AllowedModules": "dashboard,expenses",
            "Status": "Active",
        },
    )
    assert created.status_code == 200

    assigned = client.post(
        "/llp-partners",
        headers=auth_headers,
        json={
            "LLPID": "LLP001",
            "Username": "testuser@zivara.local",
            "Role": "viewer",
            "AllowedModules": "dashboard,expenses",
            "Status": "Active",
        },
    )
    assert assigned.status_code == 200

    login = client.post("/auth/login", json={"username": "testuser@zivara.local", "password": "ChangeMe123!"})
    assert login.status_code == 200
    body = login.json()
    assert body["role"] == "viewer"
    assert body["llps"][0]["llpId"] == "LLP001"


def test_bill_upload_requires_auth(client):
    res = client.post(
        "/uploads/bills",
        files={"file": ("bill.pdf", b"fake pdf", "application/pdf")},
        data={"source_type": "expense", "source_id": "EXP1"},
        headers={"X-LLP-ID": "LLP001"},
    )
    assert res.status_code == 401


def test_bill_upload_rejects_unsupported_extension(client, auth_headers):
    res = client.post(
        "/uploads/bills",
        files={"file": ("bill.exe", b"nope", "application/octet-stream")},
        data={"source_type": "expense", "source_id": "EXP1"},
        headers=auth_headers,
    )
    assert res.status_code == 400


def test_bill_upload_saves_metadata_and_updates_expense(client, auth_headers):
    created = client.post(
        "/expenses",
        headers=auth_headers,
        json={"Date": "2026-06-10", "Amount": "250", "PaidBy": "Dinu", "Description": "Hotel"},
    ).json()
    expense_id = created["data"]["ExpenseID"]
    res = client.post(
        "/uploads/bills",
        files={"file": ("bill.pdf", b"fake pdf", "application/pdf")},
        data={"source_type": "expense", "source_id": expense_id},
        headers=auth_headers,
    )
    assert res.status_code == 200
    body = res.json()
    assert body["fileId"]
    assert body["sourceType"] == "expense"
    assert body["sourceId"] == expense_id

    db = SessionLocal()
    try:
        upload = db.get(UploadedBill, body["fileId"])
        expense = db.get(Expense, expense_id)
        assert upload is not None
        assert upload.llp_id == "LLP001"
        assert upload.original_filename == "bill.pdf"
        assert expense.bill_available is True
        assert expense.bill_link
    finally:
        db.close()


def test_accounts_workbook_imports_expenses_payables_and_bank_statement(client, auth_headers):
    wb = Workbook()
    expenses = wb.active
    expenses.title = "Expenses"
    expenses.append(["LLPID", "Date", "ExpenseType", "Amount", "PaidBy", "Description"])
    expenses.append(["LLP001", "2026-06-10", "Travel", 500, "Dinu", "Cab"])

    payables = wb.create_sheet("Payables")
    payables.append(["LLPID", "VendorName", "BillNo", "BillDate", "TaxableAmount", "GSTAmount", "TDSRate"])
    payables.append(["LLP001", "ABC Co", "INV-100", "2026-06-10", 1000, 180, 10])

    bank = wb.create_sheet("BankStatement")
    bank.append(["LLPID", "Date", "Narration", "Debit", "Credit", "Balance", "ReferenceNo"])
    bank.append(["LLP001", "2026-06-11", "Vendor payment", 750, "", 9250, "UTR1"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    res = client.post(
        "/imports/accounts-workbook",
        headers=auth_headers,
        files={"file": ("accounts.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    summary = res.json()["summary"]
    assert summary["Expenses"]["imported"] == 1
    assert summary["LLPPayables"]["imported"] == 1
    assert summary["BankStatement"]["imported"] == 1

    db = SessionLocal()
    try:
        assert db.query(Expense).count() == 1
        assert db.query(LLPPayable).count() == 1
        assert db.query(CashBookEntry).count() == 1
    finally:
        db.close()


def test_neo_revenue_workbook_imports_gross_revenue_statement(client, auth_headers):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append([
        "PAN",
        "Client Name",
        "Partner",
        "Date",
        "Product",
        "Tnx Type",
        "Scheme Name",
        "Amount",
        "Commission %",
        "Remarks",
        "Gross Revenue April'26",
        "Income Type",
    ])
    sheet.append([
        "AACPH9029C",
        "SUNIL HARIDASS",
        "Manugopal A K",
        46142,
        "PMS",
        "Purchase",
        "Neo Multi-Asset Moderate Strategy",
        0,
        0.9,
        "",
        430978.662,
        "ARR",
    ])
    sheet.append([None, None, None, None, None, "Total", None, None, None, None, 430978.662, None])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    res = client.post(
        "/imports/neo-revenue-workbook",
        headers=auth_headers,
        data={"revenue_month": "Apr-2026"},
        files={"file": ("neo-gross.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    summary = res.json()["summary"]
    assert summary["NeoRevenue"]["imported"] == 1
    assert summary["NeoRevenue"]["skipped"] == 0

    db = SessionLocal()
    try:
        row = db.query(NeoRevenue).one()
        assert row.client_name == "SUNIL HARIDASS"
        assert row.partner_name == "Manugopal A K"
        assert row.llp_name == "Zivara Family Office LLP"
        assert row.revenue_month == "Apr-2026"
        assert row.revenue_amount == Decimal("430978.66")
        assert row.transaction_date.isoformat() == "2026-04-30"
        assert row.income_type == "ARR"
    finally:
        db.close()


def test_neo_revenue_workbook_imports_negative_gross_revenue(client, auth_headers):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append([
        "PAN",
        "Client Name",
        "Partner",
        "Date",
        "Product",
        "Tnx Type",
        "Scheme Name",
        "Amount",
        "Gross Revenue Aug'25",
        "Income Type",
    ])
    sheet.append(["POSPAN1234A", "Positive Client", "Manugopal A K", "2025-08-31", "PMS", "Purchase", "Positive Scheme", 1000, 500.25, "TRB"])
    sheet.append(["NEGPAN1234A", "Negative Client", "Manugopal A K", "2025-08-31", "PMS", "Reversal", "Negative Scheme", 1000, -125.75, "TRB"])
    sheet.append(["NEGPAN1234A", "Negative Client", "Manugopal A K", "2025-08-31", "PMS", "Reversal", "Negative Scheme", 1000, -125.75, "TRB"])
    sheet.append([None, None, "Total", None, None, None, None, None, 248.75, None])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    res = client.post(
        "/imports/neo-revenue-workbook",
        headers=auth_headers,
        data={"revenue_month": "Aug-2025"},
        files={"file": ("neo-negative.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    summary = res.json()["summary"]["NeoRevenue"]
    assert summary["imported"] == 3
    assert summary["skipped"] == 0

    db = SessionLocal()
    try:
        rows = db.query(NeoRevenue).all()
        assert sum(row.revenue_amount for row in rows) == Decimal("248.75")
        assert sum(1 for row in rows if row.pan == "NEGPAN1234A") == 2
    finally:
        db.close()


def test_neo_revenue_workbook_allows_small_total_rounding_difference(client, auth_headers):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append(["PAN", "Client Name", "Partner", "Date", "Product", "Tnx Type", "Scheme Name", "Amount", "Gross Revenue May'25", "Income Type"])
    sheet.append(["ROUND001A", "Rounding Client", "Manugopal A K", "2025-05-31", "PMS", "Purchase", "Rounded Scheme", 1000, 5267003.80, "TRB"])
    sheet.append([None, None, "Total", None, None, None, None, None, 5267003.78, None])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    res = client.post(
        "/imports/neo-revenue-workbook",
        headers=auth_headers,
        data={"revenue_month": "May-2025"},
        files={"file": ("neo-rounding.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    assert res.json()["summary"]["NeoRevenue"]["imported"] == 1


def test_neo_revenue_workbook_rejects_large_total_mismatch(client, auth_headers):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append(["PAN", "Client Name", "Partner", "Date", "Product", "Tnx Type", "Scheme Name", "Amount", "Gross Revenue May'25", "Income Type"])
    sheet.append(["MISMATCH1A", "Mismatch Client", "Manugopal A K", "2025-05-31", "PMS", "Purchase", "Mismatch Scheme", 1000, 1000, "TRB"])
    sheet.append([None, None, "Total", None, None, None, None, None, 999, None])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    res = client.post(
        "/imports/neo-revenue-workbook",
        headers=auth_headers,
        data={"revenue_month": "May-2025"},
        files={"file": ("neo-mismatch.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 400
    assert "Neo revenue total mismatch" in res.json()["detail"]


def test_neo_revenue_workbook_imports_only_selected_month(client, auth_headers):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append([
        "PAN",
        "Client Name",
        "Partner",
        "Date",
        "Product",
        "Tnx Type",
        "Scheme Name",
        "Amount",
        "Gross Revenue May'25",
        "Gross Revenue June'25",
        "Income Type",
    ])
    sheet.append(["MAYPAN1234A", "May Client", "Manugopal A K", "2025-05-31", "PMS", "Purchase", "May Scheme", 1000, 100, 0, "TRB"])
    sheet.append(["JUNPAN1234A", "June Client", "Manugopal A K", "2025-06-30", "PMS", "Purchase", "June Scheme", 1000, 0, 250, "TRB"])
    sheet.append([None, None, "Total", None, None, None, None, None, 100, 250, None])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    res = client.post(
        "/imports/neo-revenue-workbook",
        headers=auth_headers,
        data={"revenue_month": "Jun-2025"},
        files={"file": ("neo-selected-month.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    summary = res.json()["summary"]["NeoRevenue"]
    assert summary["imported"] == 1
    assert summary["skipped"] == 0

    db = SessionLocal()
    try:
        rows = db.query(NeoRevenue).all()
        assert len(rows) == 1
        assert rows[0].revenue_month == "Jun-2025"
        assert rows[0].pan == "JUNPAN1234A"
        assert rows[0].revenue_amount == Decimal("250.00")
    finally:
        db.close()


def test_neo_revenue_workbook_uses_selected_year_when_header_year_is_wrong(client, auth_headers):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append(["PAN", "Client Name", "Partner", "Date", "Product", "Tnx Type", "Scheme Name", "Amount", "Gross Revenue May'25", "Income Type"])
    sheet.append(["MAY26PAN1A", "May 2026 Client", "Manugopal A K", "2026-05-31", "AIF", "Purchase", "May 2026 Scheme", 1000, 1234.56, "TRB"])
    sheet.append([None, None, "Total", None, None, None, None, None, 1234.56, None])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    res = client.post(
        "/imports/neo-revenue-workbook",
        headers=auth_headers,
        data={"revenue_month": "May-2026"},
        files={"file": ("neo-may26-wrong-header-year.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    assert res.json()["summary"]["NeoRevenue"]["imported"] == 1

    db = SessionLocal()
    try:
        row = db.query(NeoRevenue).one()
        assert row.revenue_month == "May-2026"
        assert row.revenue_amount == Decimal("1234.56")
    finally:
        db.close()


def test_neo_invoices_csv_import(client, auth_headers):
    csv_data = (
        "NeoInvoiceID,RaisedBy,InvoiceType,GSTMode,IsProforma,InvoiceTitle,InvoiceNo,InvoiceDate,BillingMonth,"
        "SellerLLPID,SellerName,SellerGSTIN,BuyerName,BuyerGSTIN,Particulars,Amount,TDSRate,TDSAmount,NetPayable,Status,CreatedAt\n"
        "NINV_TEST,Manugopal,Professional Fees,With GST,No,Tax Invoice,MAK999,2026-06-13,Jun-2026,"
        "LLP001,Zivara Family Office LLP,29AAEFZ3224B1ZD,Neo Wealth Management Private Limited,27AAHCN8058K1ZV,Fees,1180,10,100,1080,Sent,2026-06-13T00:00:00Z\n"
    )
    res = client.post(
        "/imports/neo-invoices-csv",
        headers=auth_headers,
        files={"file": ("NeoInvoices.csv", csv_data.encode("utf-8"), "text/csv")},
    )
    assert res.status_code == 200
    assert res.json()["summary"]["NeoInvoices"]["imported"] == 1

    rows = client.get("/neo-invoices", headers=auth_headers).json()["data"]
    assert rows[0]["NeoInvoiceID"] == "NINV_TEST"
    assert rows[0]["InvoiceNo"] == "MAK999"
