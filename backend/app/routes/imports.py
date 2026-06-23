import csv
import re
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import current_user
from app.models import (
    BankAccount,
    CashBookEntry,
    Client,
    Expense,
    LLP,
    LLPPartner,
    LLPPayable,
    NeoInvoice,
    NeoRevenue,
    Partner,
    Receipt,
    Setting,
    User,
    Vendor,
)
from app.security import hash_password
from app.services.common import audit, make_id, normalize_key, parse_date
from app.services.neo_invoices import apply_neo_invoice_payload, create_neo_invoice
from app.services.neo_revenue import apply_client_payload, apply_revenue_payload, create_revenue, duplicate_key, serialize_revenue
from app.services.payables import calculate_amounts, payable_status

router = APIRouter(prefix="/imports", tags=["imports"])


def _text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def _value(row, *keys):
    for key in keys:
        if key in row:
            return row.get(key)
    wanted = {normalize_key(key) for key in keys}
    for key, value in row.items():
        if normalize_key(key) in wanted:
            return value
    return None


def _money(value):
    if value in (None, ""):
        return Decimal("0.00")
    try:
        return Decimal(str(value).replace(",", "").strip() or "0").quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _bool(value):
    return not _text(value).lower().startswith(("n", "false", "0", "inactive"))


def _rows(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    headers = [_text(cell.value) for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        item = {}
        for header, value in zip(headers, values):
            if header:
                item[header] = value
        if any(_text(v) for v in item.values()):
            rows.append(item)
    return rows


def _rows_any(workbook, *sheet_names):
    for sheet_name in sheet_names:
        rows = _rows(workbook, sheet_name)
        if rows:
            return rows
    return []


def _compact_key(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _row_value(row, *keys):
    wanted = {_compact_key(key) for key in keys}
    for key, value in row.items():
        if _compact_key(key) in wanted:
            return value
    return None


MONTH_ALIASES = {
    "jan": "Jan", "january": "Jan",
    "feb": "Feb", "february": "Feb",
    "mar": "Mar", "march": "Mar",
    "apr": "Apr", "april": "Apr",
    "may": "May",
    "jun": "Jun", "june": "Jun",
    "jul": "Jul", "july": "Jul",
    "aug": "Aug", "august": "Aug",
    "sep": "Sep", "sept": "Sep", "september": "Sep",
    "oct": "Oct", "october": "Oct",
    "nov": "Nov", "november": "Nov",
    "dec": "Dec", "december": "Dec",
}
NEO_REVENUE_TOTAL_TOLERANCE = Decimal("0.05")


def _revenue_month_from_header(header):
    text = str(header or "")
    month_match = re.search(
        r"\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t|tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b",
        text,
        re.IGNORECASE,
    )
    if not month_match:
        return ""
    year_match = re.search(r"(?:'|\b)(\d{2,4})\b", text[month_match.end():])
    if not year_match:
        return ""
    year = int(year_match.group(1))
    if year < 100:
        year += 2000
    month = MONTH_ALIASES[month_match.group(1).lower()]
    return f"{month}-{year}"



def _is_total_row(row):
    values = [_text(value).lower() for value in row.values()]

    return any(
        value in {"total", "grand total"}
        or value.startswith("total ")
        or value.endswith(" total")
        for value in values
    )


def _neo_revenue_total_checks(workbook, revenue_month_filter=""):
    checks = []

    for sheet in workbook.worksheets:
        headers = [_text(cell.value) for cell in sheet[1]]

        gross_headers = [
            header
            for header in headers
            if "gross" in _compact_key(header) and "revenue" in _compact_key(header)
        ]

        for gross_header in gross_headers:
            revenue_month = _revenue_month_from_header(gross_header)

            if not revenue_month:
                continue
            if revenue_month_filter and revenue_month != revenue_month_filter:
                continue

            detail_total = Decimal("0.00")
            declared_total = None
            detail_rows = 0

            for values in sheet.iter_rows(min_row=2, values_only=True):
                source = {
                    header: value
                    for header, value in zip(headers, values)
                    if header
                }

                if not any(_text(v) for v in source.values()):
                    continue

                amount = _money(source.get(gross_header))

                if _is_total_row(source):
                    declared_total = amount
                    continue

                pan = _row_value(source, "PAN")
                client_name = _row_value(source, "ClientName", "Client Name")

                if not _text(pan) and not _text(client_name):
                    continue

                if amount == Decimal("0.00"):
                    continue

                detail_total += amount
                detail_rows += 1

            checks.append({
                "sheet": sheet.title,
                "gross_header": gross_header,
                "revenue_month": revenue_month,
                "detail_total": detail_total.quantize(Decimal("0.01")),
                "declared_total": declared_total.quantize(Decimal("0.01")) if declared_total is not None else None,
                "detail_rows": detail_rows,
            })

    return checks


def _validate_neo_revenue_totals(workbook, revenue_month_filter=""):
    checks = _neo_revenue_total_checks(workbook, revenue_month_filter)

    for check in checks:
        declared_total = check["declared_total"]

        # If Neo file has no Total row, do not block import.
        # But if Total row exists, it must match.
        if declared_total is None:
            continue

        detail_total = check["detail_total"]

        if abs(detail_total - declared_total) > NEO_REVENUE_TOTAL_TOLERANCE:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Neo revenue total mismatch. "
                    f"Sheet: {check['sheet']}. "
                    f"Month: {check['revenue_month']}. "
                    f"Detail rows total: {detail_total}. "
                    f"Neo Total row: {declared_total}."
                ),
            )


def _neo_revenue_import_rows(workbook, revenue_month_filter=""):
    explicit = _rows(workbook, "NeoRevenue")

    if explicit:
        for index, row in enumerate(explicit, start=2):
            if _is_total_row(row):
                continue
            if revenue_month_filter and _text(row.get("RevenueMonth")) != revenue_month_filter:
                continue
            yield index, row
        return

    for sheet in workbook.worksheets:
        headers = [_text(cell.value) for cell in sheet[1]]

        gross_headers = [
            header
            for header in headers
            if "gross" in _compact_key(header) and "revenue" in _compact_key(header)
        ]

        for gross_header in gross_headers:
            revenue_month = _revenue_month_from_header(gross_header)

            if not revenue_month:
                continue
            if revenue_month_filter and revenue_month != revenue_month_filter:
                continue

            for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                source = {
                    header: value
                    for header, value in zip(headers, values)
                    if header
                }

                if not any(_text(v) for v in source.values()):
                    continue

                # Neo file has final Total row.
                # Use Total row for validation only, not import.
                if _is_total_row(source):
                    continue

                pan = _row_value(source, "PAN")
                client_name = _row_value(source, "ClientName", "Client Name")
                revenue_amount = source.get(gross_header)

                if not _text(pan) and not _text(client_name):
                    continue

                if _money(revenue_amount) == Decimal("0.00"):
                    continue

                row = {
                    "PAN": pan,
                    "ClientName": client_name,
                    "RMName": _row_value(source, "RMName", "RM Name", "Partner"),
                    "PartnerName": _row_value(source, "PartnerName", "Partner Name", "Partner"),
                    "TransactionDate": _row_value(source, "TransactionDate", "Transaction Date", "Date"),
                    "Product": _row_value(source, "Product"),
                    "TransactionType": _row_value(
                        source,
                        "TransactionType",
                        "Transaction Type",
                        "Tnx Type",
                        "Txn Type",
                        "Trn Type",
                    ),
                    "SchemeName": _row_value(source, "SchemeName", "Scheme Name"),
                    "InvestmentAmount": _row_value(source, "InvestmentAmount", "Investment Amount", "Amount"),
                    "CommissionPercent": _row_value(source, "CommissionPercent", "Commission %", "Commission"),
                    "Notes": _row_value(source, "Notes", "Remarks"),
                    "IncomeType": _row_value(source, "IncomeType", "Income Type") or "ARR",
                    "RevenueMonth": revenue_month,
                    "RevenueAmount": revenue_amount,
                    "StatementRef": sheet.title,
                }

                yield index, row


def _summary():
    return {"imported": 0, "skipped": 0, "errors": []}


def _safe_row(row):
    if not isinstance(row, dict):
        return {}
    return {str(key): _text(value) for key, value in row.items()}


def _skip(summary, row_no, message, row=None):
    summary["skipped"] += 1
    error = {"row": row_no, "message": message}
    if row is not None:
        error["data"] = _safe_row(row)
    summary["errors"].append(error)


def _llp_by_key(db, value):
    key = normalize_key(value)
    if not key:
        return None
    return next(
        (
            r
            for r in db.query(LLP).all()
            if normalize_key(r.id) == key
            or normalize_key(r.llp_name) == key
            or normalize_key(r.short_code) == key
        ),
        None,
    )


def _llp_id(db, row):
    llp = (
        _llp_by_key(db, _value(row, "LLPID", "LLP ID", "llpId"))
        or _llp_by_key(db, _value(row, "LLPName", "LLP Name", "llpName"))
        or _llp_by_key(db, _value(row, "ShortCode", "Short Code", "shortCode"))
    )
    return llp.id if llp else None


def _single_llp_name(db):
    llps = db.query(LLP).all()
    return llps[0].llp_name if len(llps) == 1 else ""


def _user_by_row(db, row):
    user_id = _text(_value(row, "UserID", "User ID"))
    username = _text(_value(row, "Username", "UserName", "User Name")).lower()
    if user_id:
        item = db.get(User, user_id)
        if item:
            return item
    if username:
        return db.query(User).filter(User.username == username).first()
    return None


def _vendor(db, llp_id, vendor_name):
    name_key = normalize_key(vendor_name)
    if not name_key:
        return None
    return next(
        (
            v
            for v in db.query(Vendor).filter((Vendor.llp_id == llp_id) | (Vendor.llp_id.is_(None))).all()
            if normalize_key(v.vendor_name) == name_key
        ),
        None,
    )


def _commit(db, summary, module, ref_id, user_email):
    try:
        audit(db, user_email, module, "import", ref_id)
        db.commit()
        summary["imported"] += 1
    except IntegrityError as exc:
        db.rollback()
        summary["skipped"] += 1
        summary["errors"].append({"row": ref_id, "message": str(exc.orig)})


def _clean_csv_row(row):
    return {str(key or "").strip(): value for key, value in row.items() if str(key or "").strip()}



@router.post("/neo-revenue-csv")
async def import_neo_revenue_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    filename = (file.filename or "").lower()

    if not filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Upload a NeoRevenue .csv file")

    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(StringIO(content))

    summary = _summary()

    existing_revenue_keys = {
        duplicate_key(serialize_revenue(r))
        for r in db.query(NeoRevenue).all()
    }

    seen_in_file = set()
    default_llp_name = _single_llp_name(db)

    for index, raw in enumerate(reader, start=2):
        row = _clean_csv_row(raw)

        if not _text(row.get("ClientName")):
            _skip(summary, index, "Missing ClientName", row)
            continue

        if not _text(row.get("RevenueMonth")):
            _skip(summary, index, "Missing RevenueMonth", row)
            continue

        if not _text(row.get("LLPName")) and default_llp_name:
            row["LLPName"] = default_llp_name

        revenue_id = _text(row.get("RevenueID"))
        item = db.get(NeoRevenue, revenue_id) if revenue_id else None

        key = duplicate_key(row)

        if not item and key in existing_revenue_keys:
            _skip(summary, index, "Duplicate NeoRevenue row already exists", row)
            continue

        if not item and key in seen_in_file:
            _skip(summary, index, "Duplicate NeoRevenue row inside uploaded file", row)
            continue

        try:
            if item:
                apply_revenue_payload(db, item, row)
                ref_id = item.id
            else:
                item = create_revenue(db, row)
                ref_id = item.id

            db.add(item)

            existing_revenue_keys.add(key)
            seen_in_file.add(key)

            _commit(db, summary, "neorevenue", ref_id, user.email)

        except Exception as exc:
            db.rollback()
            _skip(summary, index, str(exc) or "NeoRevenue import error", row)

    return {
        "ok": True,
        "message": "NeoRevenue CSV import complete",
        "summary": {
            "NeoRevenue": summary
        },
    }


@router.post("/neo-invoices-csv")
async def import_neo_invoices_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    filename = (file.filename or "").lower()
    if not filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Upload a NeoInvoices .csv file")

    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(StringIO(content))
    summary = _summary()
    for index, raw in enumerate(reader, start=2):
        row = _clean_csv_row(raw)
        invoice_id = _text(row.get("NeoInvoiceID"))
        invoice_no = _text(row.get("InvoiceNo"))
        billing_month = _text(row.get("BillingMonth"))
        llp_id = _llp_id(db, row) or _text(row.get("SellerLLPID")) or None
        if not invoice_no or not billing_month:
            _skip(summary, index, "Missing InvoiceNo or BillingMonth", row)
            continue
        item = db.get(NeoInvoice, invoice_id) if invoice_id else None
        if not item:
            item = db.query(NeoInvoice).filter(NeoInvoice.llp_id == llp_id, NeoInvoice.invoice_no == invoice_no).first()
        if item:
            apply_neo_invoice_payload(db, item, row, llp_id or item.llp_id)
            ref_id = item.id
        else:
            item = create_neo_invoice(db, row, llp_id)
            ref_id = item.id
        _commit(db, summary, "neoinvoices", ref_id, user.email)

    return {"ok": True, "message": "NeoInvoices import complete", "summary": {"NeoInvoices": summary}}


@router.post("/neo-revenue-workbook")
async def import_neo_revenue_workbook(
    revenue_month: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    filename = (file.filename or "").lower()
    selected_month = _text(revenue_month)

    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload an Excel .xlsx file")
    if not selected_month:
        raise HTTPException(status_code=400, detail="Revenue month is required")

    content = await file.read()
    workbook = load_workbook(BytesIO(content), data_only=True)
    _validate_neo_revenue_totals(workbook, selected_month)

    summary = _summary()
    database_revenue_keys = {duplicate_key(serialize_revenue(r)) for r in db.query(NeoRevenue).all()}
    default_llp_name = _single_llp_name(db)
    matched_rows = 0

    for index, row in _neo_revenue_import_rows(workbook, selected_month):
        matched_rows += 1
        if not _text(row.get("ClientName")) or not _text(row.get("RevenueMonth")):
            _skip(summary, index, "Missing ClientName or RevenueMonth", row)
            continue
        if not _text(row.get("LLPName")) and default_llp_name:
            row["LLPName"] = default_llp_name
        item = db.get(NeoRevenue, _text(row.get("RevenueID"))) if _text(row.get("RevenueID")) else None
        key = duplicate_key(row)
        if not item and key in database_revenue_keys:
            _skip(summary, index, "Duplicate NeoRevenue row already exists in database", row)
            continue
        if item:
            apply_revenue_payload(db, item, row)
        else:
            item = create_revenue(db, row)
        db.add(item)
        _commit(db, summary, "neorevenue", item.id, user.email)

    if matched_rows == 0:
        _skip(summary, selected_month, f"No Neo Revenue rows found for {selected_month}")

    return {
        "ok": True,
        "message": f"NeoRevenue import complete for {selected_month}",
        "summary": {"NeoRevenue": summary},
    }


@router.post("/accounts-workbook")
async def import_accounts_workbook(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload an Excel .xlsx file")

    content = await file.read()
    workbook = load_workbook(BytesIO(content), data_only=True)
    _validate_neo_revenue_totals(workbook)
    result = {name: _summary() for name in [
        "Settings",
        "LLPs",
        "Users",
        "Clients",
        "NeoRevenue",
        "Partners",
        "LLPPartners",
        "Vendors",
        "BankAccounts",
        "LLPPayables",
        "Expenses",
        "Receipts",
        "CashBook",
        "BankStatement",
    ]}

    for index, row in enumerate(_rows(workbook, "Settings"), start=2):
        key = _text(row.get("Key"))
        if not key:
            _skip(result["Settings"], index, "Missing Key", row)
            continue
        db.merge(Setting(id=make_id("SET"), key=key, value=_text(row.get("Value"))))
        _commit(db, result["Settings"], "settings", key, user.email)

    for index, row in enumerate(_rows(workbook, "LLPs"), start=2):
        llp_id = _text(row.get("LLPID")) or make_id("LLP")
        if not _text(row.get("LLPName")):
            _skip(result["LLPs"], index, "Missing LLPName", row)
            continue
        existing = db.get(LLP, llp_id)
        item = existing or LLP(id=llp_id, llp_name="")
        item.llp_name = _text(row.get("LLPName"))
        item.short_code = _text(row.get("ShortCode")) or llp_id
        item.gstin = _text(row.get("GSTIN"))
        item.pan = _text(row.get("PAN"))
        item.address = _text(row.get("Address"))
        item.status = _text(row.get("Status")) or "Active"
        db.add(item)
        _commit(db, result["LLPs"], "llps", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "Users"), start=2):
        username = _text(row.get("Username")).lower()
        if not username:
            _skip(result["Users"], index, "Missing Username", row)
            continue
        item = db.query(User).filter(User.username == username).first() or User(
            id=_text(row.get("UserID")) or make_id("USR"),
            username=username,
            email=username,
            password_hash=hash_password("ChangeMe123!"),
        )
        item.name = _text(row.get("Name")) or username
        item.email = _text(row.get("Email")) or username
        item.role = _text(row.get("Role")) or "viewer"
        item.allowed_modules = _text(row.get("AllowedModules"))
        item.status = _text(row.get("Status")) or "Active"
        db.add(item)
        _commit(db, result["Users"], "users", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "Partners"), start=2):
        item = db.get(Partner, _text(row.get("PartnerID"))) if _text(row.get("PartnerID")) else None
        item = item or Partner(id=_text(row.get("PartnerID")) or make_id("PTR"), partner_name="")
        item.partner_name = _text(row.get("PartnerName"))
        item.llp_id = _llp_id(db, row)
        item.email = _text(row.get("Email"))
        item.mobile = _text(row.get("Mobile"))
        item.status = _text(row.get("Status")) or "Active"
        db.add(item)
        _commit(db, result["Partners"], "partners", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "Clients"), start=2):
        if not _text(row.get("ClientName")):
            _skip(result["Clients"], index, "Missing ClientName", row)
            continue
        item = db.get(Client, _text(row.get("ClientID"))) if _text(row.get("ClientID")) else None
        item = item or Client(id=_text(row.get("ClientID")) or make_id("CLT"), client_name="")
        apply_client_payload(item, row)
        db.add(item)
        _commit(db, result["Clients"], "clients", item.id, user.email)

    database_revenue_keys = {duplicate_key(serialize_revenue(r)) for r in db.query(NeoRevenue).all()}
    default_llp_name = _single_llp_name(db)
    for index, row in _neo_revenue_import_rows(workbook):
        if not _text(row.get("ClientName")) or not _text(row.get("RevenueMonth")):
            _skip(result["NeoRevenue"], index, "Missing ClientName or RevenueMonth", row)
            continue
        if not _text(row.get("LLPName")) and default_llp_name:
            row["LLPName"] = default_llp_name
        item = db.get(NeoRevenue, _text(row.get("RevenueID"))) if _text(row.get("RevenueID")) else None
        key = duplicate_key(row)
        if not item and key in database_revenue_keys:
            _skip(result["NeoRevenue"], index, "Duplicate NeoRevenue row already exists in database", row)
            continue
        if item:
            apply_revenue_payload(db, item, row)
        else:
            item = create_revenue(db, row)
        db.add(item)
        _commit(db, result["NeoRevenue"], "neorevenue", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "LLPPartners"), start=2):
        llp_id = _llp_id(db, row)
        target = _user_by_row(db, row)
        if not llp_id or not target:
            _skip(result["LLPPartners"], index, "Missing matching LLP or Username", row)
            continue
        item = db.get(LLPPartner, _text(row.get("MappingID"))) if _text(row.get("MappingID")) else None
        item = item or LLPPartner(id=_text(row.get("MappingID")) or make_id("MAP"), llp_id=llp_id, user_id=target.id)
        item.role = _text(row.get("Role")) or "partner"
        item.percentage = _text(row.get("Percentage"))
        item.allowed_modules = _text(row.get("AllowedModules"))
        item.status = _text(row.get("Status")) or "Active"
        db.add(item)
        _commit(db, result["LLPPartners"], "llp_partners", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "Vendors"), start=2):
        item = db.get(Vendor, _text(row.get("VendorID"))) if _text(row.get("VendorID")) else None
        item = item or Vendor(id=_text(row.get("VendorID")) or make_id("VND"), vendor_name="")
        item.vendor_name = _text(row.get("VendorName"))
        item.llp_id = _llp_id(db, row)
        item.category = _text(row.get("Category"))
        item.gstin = _text(row.get("GSTIN"))
        item.pan = _text(row.get("PAN"))
        item.state = _text(row.get("State"))
        item.notes = _text(row.get("Notes"))
        item.status = _text(row.get("Status")) or "Active"
        db.add(item)
        _commit(db, result["Vendors"], "vendors", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "BankAccounts"), start=2):
        llp_id = _llp_id(db, row)
        if not llp_id:
            _skip(result["BankAccounts"], index, "Missing matching LLP", row)
            continue
        item = db.get(BankAccount, _text(row.get("AccountID"))) if _text(row.get("AccountID")) else None
        item = item or BankAccount(id=_text(row.get("AccountID")) or make_id("BANK"), llp_id=llp_id, account_name="", bank_name="", account_number="")
        item.llp_id = llp_id
        item.account_name = _text(row.get("AccountName"))
        item.bank_name = _text(row.get("BankName"))
        item.account_number = _text(row.get("AccountNumber"))
        item.ifsc = _text(row.get("IFSC"))
        item.account_type = _text(row.get("AccountType")) or "Current"
        item.branch = _text(row.get("Branch"))
        item.opening_balance = _money(row.get("OpeningBalance"))
        item.current_balance = _money(row.get("CurrentBalance")) or item.opening_balance
        item.is_active = _bool(row.get("IsActive"))
        item.notes = _text(row.get("Notes"))
        db.add(item)
        _commit(db, result["BankAccounts"], "bankaccounts", item.id, user.email)

    for index, row in enumerate(_rows_any(workbook, "LLPPayables", "Payables"), start=2):
        llp_id = _llp_id(db, row)
        if not llp_id:
            _skip(result["LLPPayables"], index, "Missing matching LLP", row)
            continue
        vendor = _vendor(db, llp_id, row.get("VendorName"))
        merged = {key: row.get(key) for key in row}
        taxable, gst, gross, tds_rate, tds, net = calculate_amounts(merged)
        item = db.get(LLPPayable, _text(row.get("PayableID"))) if _text(row.get("PayableID")) else None
        item = item or LLPPayable(id=_text(row.get("PayableID")) or make_id("PAY"), llp_id=llp_id, vendor_name="", bill_no="", normalized_bill_no="")
        item.llp_id = llp_id
        item.vendor_id = vendor.id if vendor else None
        item.vendor_name = _text(row.get("VendorName"))
        item.vendor_category = _text(row.get("VendorCategory"))
        item.vendor_gstin = _text(row.get("VendorGSTIN"))
        item.vendor_pan = _text(row.get("VendorPAN"))
        item.bill_no = _text(row.get("BillNo"))
        item.normalized_bill_no = normalize_key(item.bill_no)
        item.bill_date = parse_date(row.get("BillDate"))
        item.due_date = parse_date(row.get("DueDate"))
        item.expense_type = _text(row.get("ExpenseType")) or "Vendor Bill"
        item.description = _text(row.get("Description"))
        item.taxable_amount, item.gst_amount, item.gross_amount = taxable, gst, gross
        item.tds_section = _text(row.get("TDSSection"))
        item.tds_rate, item.tds_amount, item.net_payable = tds_rate, tds, net
        item.paid_amount = _money(row.get("PaidAmount"))
        item.payment_date = parse_date(row.get("PaymentDate"))
        item.payment_mode = _text(row.get("PaymentMode")) or "Bank"
        item.bank_account = _text(row.get("BankAccount"))
        item.reference_no = _text(row.get("ReferenceNo"))
        item.challan_no = _text(row.get("ChallanNo"))
        item.challan_date = parse_date(row.get("ChallanDate"))
        item.interest_amount = _money(row.get("InterestAmount"))
        item.status = payable_status(item.net_payable, item.paid_amount, _text(row.get("Status")))
        item.notes = _text(row.get("Notes"))
        db.add(item)
        _commit(db, result["LLPPayables"], "payables", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "Expenses"), start=2):
        llp_id = _llp_id(db, row)
        if not llp_id:
            _skip(result["Expenses"], index, "Missing matching LLP", row)
            continue
        item = db.get(Expense, _text(row.get("ExpenseID"))) if _text(row.get("ExpenseID")) else None
        item = item or Expense(id=_text(row.get("ExpenseID")) or make_id("EXP"), llp_id=llp_id)
        item.llp_id = llp_id
        item.expense_date = parse_date(row.get("Date"))
        item.expense_type = _text(row.get("ExpenseType")) or "Misc"
        item.category = _text(row.get("Category"))
        item.paid_by = _text(row.get("PaidBy"))
        item.charge_to = _text(row.get("ChargeTo"))
        item.reimburse_to = _text(row.get("ReimburseTo"))
        item.payment_mode = _text(row.get("PaymentMode")) or "Cash"
        item.taxable_value = _money(row.get("TaxableValue"))
        item.cgst_amount = _money(row.get("CGSTAmount"))
        item.sgst_amount = _money(row.get("SGSTAmount"))
        item.igst_amount = _money(row.get("IGSTAmount"))
        item.gst_amount = _money(row.get("GSTAmount"))
        item.amount = _money(row.get("Amount"))
        item.vendor_or_person = _text(row.get("VendorOrPerson"))
        item.description = _text(row.get("Description"))
        item.bill_available = _bool(row.get("BillAvailable"))
        item.bill_link = _text(row.get("BillLink"))
        item.employee_name = _text(row.get("EmployeeName"))
        item.billing_month = _text(row.get("BillingMonth"))
        item.travel_id = _text(row.get("TravelID"))
        item.partner_allocations = _text(row.get("PartnerAllocations"))
        item.notes = _text(row.get("Notes"))
        item.status = _text(row.get("Status")) or "Draft"
        item.reimburse_mode = _text(row.get("ReimburseMode"))
        item.reimburse_account = _text(row.get("ReimburseAccount"))
        item.reimburse_date = parse_date(row.get("ReimburseDate"))
        item.reimburse_ref = _text(row.get("ReimburseRef"))
        item.reimburse_by = _text(row.get("ReimburseBy"))
        item.approved_by = _text(row.get("ApprovedBy"))
        db.add(item)
        _commit(db, result["Expenses"], "expenses", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "Receipts"), start=2):
        llp_id = _llp_id(db, row)
        if not llp_id:
            _skip(result["Receipts"], index, "Missing matching LLP", row)
            continue
        item = db.get(Receipt, _text(row.get("ReceiptID"))) if _text(row.get("ReceiptID")) else None
        item = item or Receipt(id=_text(row.get("ReceiptID")) or make_id("REC"), llp_id=llp_id)
        item.llp_id = llp_id
        item.receipt_date = parse_date(row.get("Date"))
        item.reference_type = _text(row.get("ReferenceType"))
        item.reference_no = _text(row.get("ReferenceNo") or row.get("InvoiceNo"))
        item.month = _text(row.get("Month"))
        item.amount_received = _money(row.get("AmountReceived"))
        item.receipt_mode = _text(row.get("ReceiptMode")) or "Bank"
        item.bank_account = _text(row.get("BankAccount"))
        item.notes = _text(row.get("Notes"))
        db.add(item)
        _commit(db, result["Receipts"], "receipts", item.id, user.email)

    for index, row in enumerate(_rows(workbook, "CashBook"), start=2):
        llp_id = _llp_id(db, row)
        if not llp_id:
            _skip(result["CashBook"], index, "Missing matching LLP", row)
            continue
        item = db.get(CashBookEntry, _text(row.get("EntryID"))) if _text(row.get("EntryID")) else None
        item = item or CashBookEntry(id=_text(row.get("EntryID")) or make_id("CASH"), llp_id=llp_id)
        item.llp_id = llp_id
        item.entry_date = parse_date(row.get("Date"))
        item.entry_type = _text(row.get("Type")) or "Payment"
        item.opening_balance = _money(row.get("OpeningBalance"))
        item.amount_in = _money(row.get("AmountIn"))
        item.amount_out = _money(row.get("AmountOut"))
        item.closing_balance = _money(row.get("ClosingBalance")) or (item.opening_balance + item.amount_in - item.amount_out)
        item.reference_type = _text(row.get("ReferenceType")) or "Manual"
        item.reference_id = _text(row.get("ReferenceID"))
        item.description = _text(row.get("Description"))
        item.paid_by = _text(row.get("PaidBy"))
        db.add(item)
        _commit(db, result["CashBook"], "cashbook", item.id, user.email)

    for index, row in enumerate(_rows_any(workbook, "BankStatement", "BankStatements"), start=2):
        llp_id = _llp_id(db, row)
        if not llp_id:
            _skip(result["BankStatement"], index, "Missing matching LLP", row)
            continue
        debit = _money(_value(row, "Debit", "Withdrawal", "AmountOut"))
        credit = _money(_value(row, "Credit", "Deposit", "AmountIn"))
        amount = _money(_value(row, "Amount"))
        if amount > 0 and not credit:
            credit = amount
        elif amount < 0 and not debit:
            debit = abs(amount)
        entry_type = _text(_value(row, "Type")) or ("Receipt" if credit > 0 else "Payment")
        opening = _money(_value(row, "OpeningBalance", "Opening Balance"))
        closing = _money(_value(row, "ClosingBalance", "Closing Balance", "Balance"))
        item = db.get(CashBookEntry, _text(_value(row, "EntryID", "TransactionID", "TxnID"))) if _text(_value(row, "EntryID", "TransactionID", "TxnID")) else None
        item = item or CashBookEntry(id=_text(_value(row, "EntryID", "TransactionID", "TxnID")) or make_id("CASH"), llp_id=llp_id)
        item.llp_id = llp_id
        item.entry_date = parse_date(_value(row, "Date", "TxnDate", "TransactionDate", "ValueDate"))
        item.entry_type = entry_type
        item.opening_balance = opening
        item.amount_in = credit
        item.amount_out = debit
        item.closing_balance = closing or (opening + credit - debit)
        item.reference_type = _text(_value(row, "ReferenceType", "RefType")) or "BankStatement"
        item.reference_id = _text(_value(row, "ReferenceID", "ReferenceNo", "UTR", "ChequeNo", "Narration"))
        item.description = _text(_value(row, "Description", "Narration", "Particulars", "Remarks"))
        item.paid_by = _text(_value(row, "PaidBy", "Payee", "Payer"))
        db.add(item)
        _commit(db, result["BankStatement"], "bankstatement", item.id, user.email)

    return {"ok": True, "message": "Workbook import complete", "summary": result}
